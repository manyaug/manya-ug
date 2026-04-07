"""
Create a sample English MCQ Excel with two sheets:
  Sheet 1 (Raw): Original questions
  Sheet 2 (Rephrased): 3 variants per raw question
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── STYLE DEFINITIONS ───
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
rephrased_fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

HEADERS = [
    'qid', 'term', 'topic', 'subtopic', 'difficulty', 'marked_ple',
    'questiontype', 'parentid', 'orderinparent', 'questiontext',
    'optiona', 'optionb', 'optionc', 'optiond', 'correctanswer',
    'hint', 'detailedsolution', 'engine_type'
]

def style_headers(ws, fill):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.freeze_panes = 'A2'
    # Auto-width
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions[get_column_letter(10)].width = 50  # questiontext
    ws.column_dimensions[get_column_letter(16)].width = 30  # hint
    ws.column_dimensions[get_column_letter(17)].width = 40  # detailedsolution

# ═══════════════════════════════════════════════════════════════
# SHEET 1: RAW QUESTIONS
# ═══════════════════════════════════════════════════════════════
ws_raw = wb.active
ws_raw.title = 'Raw'
style_headers(ws_raw, header_fill)

raw_questions = [
    # ── Quest 01: Holiday Kickoff (Vocabulary + Parts of Speech) ──
    ['ENG-T1-01-001', 'T1', 'quest_01_holiday_kickoff', 'Vocabulary', 'E', 'no', 'MCQ', '', 1,
     'What does the word "commence" mean?',
     'To begin', 'To finish', 'To travel', 'To sleep', 'To begin',
     'Think about the start of something', '"Commence" means to start or begin officially.', 'MCQ'],
    
    ['ENG-T1-01-002', 'T1', 'quest_01_holiday_kickoff', 'Vocabulary', 'E', 'no', 'MCQ', '', 2,
     'Which word means "a quiet countryside area"?',
     'Urban', 'Rural', 'Suburb', 'Metro', 'Rural',
     'Opposite of city', '"Rural" means the quiet countryside or village area.', 'MCQ'],
    
    ['ENG-T1-01-003', 'T1', 'quest_01_holiday_kickoff', 'Parts_of_Speech', 'E', 'no', 'MCQ', '', 3,
     'Which of these is a NOUN?',
     'Run', 'Farm', 'Quickly', 'Beautiful', 'Farm',
     'A noun is a person, place, or thing', '"Farm" is a place, so it is a noun.', 'MCQ'],
    
    ['ENG-T1-01-004', 'T1', 'quest_01_holiday_kickoff', 'Parts_of_Speech', 'M', 'yes', 'MCQ', '', 4,
     'Identify the VERB in: "The children played in the field."',
     'children', 'played', 'field', 'the', 'played',
     'A verb is an action word', '"Played" is the action done by the children.', 'MCQ'],
    
    ['ENG-T1-01-005', 'T1', 'quest_01_holiday_kickoff', 'Vocabulary', 'M', 'no', 'MCQ', '', 5,
     'What does "fare" mean in: "The bus fare is 5000 shillings"?',
     'A type of bus', 'Money paid for transport', 'A long journey', 'Food on a bus', 'Money paid for transport',
     'Think about what you pay for a ride', '"Fare" is the money you pay for transport.', 'MCQ'],
    
    # ── Quest 02: Going To Mastery (Future Plans) ──
    ['ENG-T1-02-001', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', '', 1,
     'Complete: She _____ going to visit her aunt.',
     'is', 'are', 'am', 'were', 'is',
     'Check the subject — "She" is singular', '"She" is a singular subject → use "is".', 'MCQ'],
    
    ['ENG-T1-02-002', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', '', 2,
     'Complete: I _____ going to read my books.',
     'am', 'is', 'are', 'was', 'am',
     'What word goes with "I"?', '"I" always takes "am".', 'MCQ'],
    
    ['ENG-T1-02-003', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', '', 3,
     'Complete: They _____ going to play football.',
     'are', 'is', 'am', 'was', 'are',
     '"They" is plural', '"They" is plural → use "are".', 'MCQ'],
    
    ['ENG-T1-02-004', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'M', 'yes', 'MCQ', '', 4,
     'Which sentence uses "going to" correctly?',
     'I am going to cook rice.', 'I going to cook rice.', 'I is going to cook rice.', 'I are going cook rice.', 'I am going to cook rice.',
     'Subject + am/is/are + going to + verb', 'The correct formula is: Subject + am/is/are + going to + verb.', 'MCQ'],
    
    ['ENG-T1-02-005', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'M', 'no', 'MCQ', '', 5,
     'Complete: Manya _____ going to visit his grandparents.',
     'is', 'am', 'are', 'be', 'is',
     '"Manya" is one person (singular)', '"Manya" is a singular noun → use "is".', 'MCQ'],
    
    # ── Quest 03: Question Tags Mastery ──
    ['ENG-T1-03-001', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'E', 'no', 'MCQ', '', 1,
     'Complete the tag: Manya is going to the village, _____?',
     "isn't he", "is he", "wasn't he", "doesn't he", "isn't he",
     'Positive statement → negative tag', 'Positive "is" → negative "isn\'t he".', 'MCQ'],
    
    ['ENG-T1-03-002', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'E', 'no', 'MCQ', '', 2,
     "Complete the tag: We won't stay in town, _____?",
     'will we', "won't we", 'do we', "don't we", 'will we',
     "Negative statement → positive tag", '"Won\'t" is negative → positive tag "will we".', 'MCQ'],
    
    ['ENG-T1-03-003', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'M', 'yes', 'MCQ', '', 3,
     'Complete the tag: I am right, _____?',
     "aren't I", "am I", "isn't I", "don't I", "aren't I",
     'Special rule for "I am"', '"I am" uses the special tag "aren\'t I".', 'MCQ'],
    
    ['ENG-T1-03-004', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'M', 'no', 'MCQ', '', 4,
     "Complete the tag: Let us go for vacation, _____?",
     'shall we', "shan't we", 'will we', "won't we", 'shall we',
     '"Let us" has a special tag', '"Let us" uses "shall we".', 'MCQ'],
    
    # ── Quest 04: Reported Speech Mastery ──
    ['ENG-T1-04-001', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'E', 'no', 'MCQ', '', 1,
     'Direct: "I am going home." Reported: He said that he _____ going home.',
     'was', 'is', 'am', 'were', 'was',
     'am → was in reported speech', '"Am" changes to "was" in reported speech.', 'MCQ'],
    
    ['ENG-T1-04-002', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'E', 'no', 'MCQ', '', 2,
     'Direct: "We will enjoy." Reported: They said they _____ enjoy.',
     'would', 'will', 'shall', 'can', 'would',
     'will → would', '"Will" changes to "would" in reported speech.', 'MCQ'],
    
    ['ENG-T1-04-003', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'M', 'yes', 'MCQ', '', 3,
     'Direct: "I can swim." Reported: Manya said that he _____ swim.',
     'could', 'can', 'was', 'is', 'could',
     'can → could', '"Can" changes to "could" in reported speech.', 'MCQ'],
    
    # ── Quest 06: Adjectives -ed/-ing ──
    ['ENG-T1-06-001', 'T1', 'quest_06_feeling_and_facts', 'Adjectives_ing_ed', 'E', 'no', 'MCQ', '', 1,
     'The movie was very _____. (It caused a feeling)',
     'exciting', 'excited', 'excite', 'excites', 'exciting',
     '-ING describes the cause', 'Use "-ing" to describe the thing that causes the feeling.', 'MCQ'],
    
    ['ENG-T1-06-002', 'T1', 'quest_06_feeling_and_facts', 'Adjectives_ing_ed', 'E', 'no', 'MCQ', '', 2,
     'Manya was very _____ by the village. (He had a feeling)',
     'excited', 'exciting', 'excite', 'excites', 'excited',
     "-ED describes the person's feeling", 'Use "-ed" to describe how the person feels.', 'MCQ'],
    
    # ── Quest 07: Conditionals Type 3 ──
    ['ENG-T1-07-001', 'T1', 'quest_07_past_regrets', 'Conditionals_Type3', 'E', 'no', 'MCQ', '', 1,
     'If I had known, I _____ have come.',
     'would', 'will', 'shall', 'can', 'would',
     'Type 3: If had + past participle → would have + past participle',
     'Type 3 conditionals use "would have" in the main clause.', 'MCQ'],
    
    ['ENG-T1-07-002', 'T1', 'quest_07_past_regrets', 'Conditionals_Type3', 'M', 'yes', 'MCQ', '', 2,
     'She would have passed if she _____ revised.',
     'had', 'has', 'have', 'was', 'had',
     'Type 3: If + had + past participle',
     'The "if" clause uses "had" + past participle.', 'MCQ'],
    
    # ── Quest 08: Active & Passive Voice ──
    ['ENG-T1-08-001', 'T1', 'quest_08_voice_mastery', 'Active_Passive', 'E', 'no', 'MCQ', '', 1,
     'Change to passive: "Polly is eating a mango."',
     'A mango is being eaten by Polly.', 'A mango was eaten by Polly.', 'Polly eats a mango.', 'The mango ate Polly.', 'A mango is being eaten by Polly.',
     'Present continuous passive: is/are being + past participle',
     'Present continuous passive: Object + is being + past participle + by Subject.', 'MCQ'],
    
    ['ENG-T1-08-002', 'T1', 'quest_08_voice_mastery', 'Active_Passive', 'M', 'yes', 'MCQ', '', 2,
     'Change to passive: "The teacher punished the boy."',
     'The boy was punished by the teacher.', 'The boy is punished by the teacher.', 'The boy punished the teacher.', 'The teacher was punished.', 'The boy was punished by the teacher.',
     'Simple past passive: was/were + past participle',
     'Simple past passive: Object + was/were + past participle + by Subject.', 'MCQ'],
]

for row_idx, row_data in enumerate(raw_questions, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_raw.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# ═══════════════════════════════════════════════════════════════
# SHEET 2: REPHRASED QUESTIONS (3 variants per raw)
# ═══════════════════════════════════════════════════════════════
ws_reph = wb.create_sheet('Rephrased')
style_headers(ws_reph, rephrased_fill)

rephrased_questions = [
    # Variants for ENG-T1-02-001 (Going To - "She _____ going to visit")
    ['ENG-T1-02-001-V1', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', 'ENG-T1-02-001', 1,
     'Complete: He _____ going to cook dinner.',
     'is', 'are', 'am', 'were', 'is',
     '"He" is singular', '"He" is a singular subject → use "is".', 'MCQ'],
    
    ['ENG-T1-02-001-V2', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', 'ENG-T1-02-001', 2,
     'Complete: Kiki _____ going to play.',
     'is', 'are', 'am', 'was', 'is',
     '"Kiki" is one person', '"Kiki" is singular → use "is".', 'MCQ'],
    
    ['ENG-T1-02-001-V3', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'M', 'no', 'MCQ', 'ENG-T1-02-001', 3,
     'Complete: The bird _____ going to fly south.',
     'is', 'are', 'am', 'were', 'is',
     '"The bird" is one thing', '"The bird" is singular → use "is".', 'MCQ'],
    
    # Variants for ENG-T1-02-003 (Going To - "They _____ going to play")
    ['ENG-T1-02-003-V1', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', 'ENG-T1-02-003', 1,
     'Complete: We _____ going to swim.',
     'are', 'is', 'am', 'was', 'are',
     '"We" is plural', '"We" is plural → use "are".', 'MCQ'],
    
    ['ENG-T1-02-003-V2', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'E', 'no', 'MCQ', 'ENG-T1-02-003', 2,
     'Complete: The students _____ going to revise.',
     'are', 'is', 'am', 'were', 'are',
     '"The students" is plural', '"The students" is plural → use "are".', 'MCQ'],
    
    ['ENG-T1-02-003-V3', 'T1', 'quest_02_going_to_mastery', 'Going_To', 'M', 'no', 'MCQ', 'ENG-T1-02-003', 3,
     'Complete: You _____ going to enjoy the village.',
     'are', 'is', 'am', 'was', 'are',
     '"You" always takes "are"', '"You" always uses "are" (even for one person).', 'MCQ'],

    # Variants for ENG-T1-03-001 (Question Tags)
    ['ENG-T1-03-001-V1', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'E', 'no', 'MCQ', 'ENG-T1-03-001', 1,
     'Complete the tag: She is a student, _____?',
     "isn't she", "is she", "doesn't she", "does she", "isn't she",
     'Positive statement → negative tag', 'Positive "is" → negative tag "isn\'t she".', 'MCQ'],
    
    ['ENG-T1-03-001-V2', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'E', 'no', 'MCQ', 'ENG-T1-03-001', 2,
     'Complete the tag: They are playing, _____?',
     "aren't they", "are they", "don't they", "do they", "aren't they",
     'Positive statement → negative tag', 'Positive "are" → negative tag "aren\'t they".', 'MCQ'],
    
    ['ENG-T1-03-001-V3', 'T1', 'quest_03_question_tags_mastery', 'Question_Tags', 'M', 'no', 'MCQ', 'ENG-T1-03-001', 3,
     'Complete the tag: The bus has arrived, _____?',
     "hasn't it", "has it", "doesn't it", "isn't it", "hasn't it",
     'Positive "has" → negative tag', 'Positive "has" → negative tag "hasn\'t it".', 'MCQ'],

    # Variants for ENG-T1-04-001 (Reported Speech)
    ['ENG-T1-04-001-V1', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'E', 'no', 'MCQ', 'ENG-T1-04-001', 1,
     'Direct: "She is cooking." Reported: He said she _____ cooking.',
     'was', 'is', 'are', 'were', 'was',
     'is → was', '"Is" changes to "was" in reported speech.', 'MCQ'],
    
    ['ENG-T1-04-001-V2', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'E', 'no', 'MCQ', 'ENG-T1-04-001', 2,
     'Direct: "I am happy." Reported: She said she _____ happy.',
     'was', 'is', 'am', 'are', 'was',
     'am → was', '"Am" changes to "was" in reported speech.', 'MCQ'],
    
    ['ENG-T1-04-001-V3', 'T1', 'quest_04_reported_speech_mastery', 'Reported_Speech', 'M', 'no', 'MCQ', 'ENG-T1-04-001', 3,
     'Direct: "They are playing." Reported: She said they _____ playing.',
     'were', 'are', 'was', 'is', 'were',
     'are → were', '"Are" changes to "were" in reported speech.', 'MCQ'],
]

for row_idx, row_data in enumerate(rephrased_questions, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_reph.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Save
output_path = r'd:\manya_app\scripts\excel_sync\english-p7-question-bank.xlsx'
wb.save(output_path)
print(f'✅ Created: {output_path}')
print(f'   Sheet 1 (Raw): {len(raw_questions)} questions')
print(f'   Sheet 2 (Rephrased): {len(rephrased_questions)} variants')
